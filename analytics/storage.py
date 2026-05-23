# analytics/storage.py
"""
Draw analytics and flexible storage format.

This module provides:
1. A pliable JSON/Pydantic format for storing draws (DrawStorage)
2. Comprehensive analytics generation (DrawAnalytics)
3. Excel export with multiple analysis sheets
4. Draw modification utilities for testing
"""

import json
import sys
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
import pandas as pd
from pydantic import BaseModel, Field

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from models import Team, Club, Grade, Roster


# ============== Pliable Draw Storage Format ==============

class StoredGame(BaseModel):
    """A single game in pliable storage format."""
    game_id: str = Field(..., description="Unique identifier for this game")
    team1: str = Field(..., description="Home team name")
    team2: str = Field(..., description="Away team name")
    grade: str = Field(..., description="Grade name")
    week: int = Field(..., description="Week number")
    round_no: int = Field(..., description="Round number")
    date: str = Field(..., description="Date string (YYYY-MM-DD)")
    day: str = Field(..., description="Day of week")
    time: str = Field(..., description="Game time (HH:MM)")
    day_slot: int = Field(..., description="Slot number within the day")
    field_name: str = Field(..., description="Field name")
    field_location: str = Field(..., description="Field location/venue")
    
    @property
    def teams(self) -> Tuple[str, str]:
        return (self.team1, self.team2)
    
    def to_key(self) -> Tuple:
        """Convert back to X-dict key format."""
        return (
            self.team1, self.team2, self.grade, self.day, self.day_slot,
            self.time, self.week, self.date, self.round_no,
            self.field_name, self.field_location
        )


class DrawStorage(BaseModel):
    """Pliable storage format for complete draws."""
    version: str = Field(default="1.0", description="Format version")
    created_at: str = Field(default_factory=lambda: datetime.now().isoformat())
    description: str = Field(default="", description="Optional description")
    num_weeks: int = Field(default=0, description="Total weeks in draw")
    num_games: int = Field(default=0, description="Total games in draw")
    games: List[StoredGame] = Field(default_factory=list)
    metadata: Dict[str, Any] = Field(default_factory=dict)
    
    @classmethod
    def from_X_solution(cls, X_solution: Dict, description: str = "") -> "DrawStorage":
        """Create DrawStorage from an X solution dictionary."""
        games = []
        game_id = 0
        
        for key, value in X_solution.items():
            # Handle both solved (has solution_value) and dict value formats
            is_scheduled = False
            if hasattr(value, 'solution_value'):
                is_scheduled = value.solution_value() > 0.5
            else:
                is_scheduled = bool(value)
            
            if not is_scheduled or len(key) < 11:
                continue
            
            t1, t2, grade, day, day_slot, time, week, date, round_no, field_name, field_location = key[:11]
            
            game = StoredGame(
                game_id=f"G{game_id:05d}",
                team1=t1,
                team2=t2,
                grade=grade,
                week=int(week),
                round_no=int(round_no),
                date=str(date),
                day=day,
                time=str(time),
                day_slot=int(day_slot),
                field_name=field_name,
                field_location=field_location
            )
            games.append(game)
            game_id += 1
        
        weeks = set(g.week for g in games)
        
        return cls(
            description=description,
            num_weeks=len(weeks),
            num_games=len(games),
            games=games
        )
    
    @classmethod
    def from_roster(cls, roster: Roster, description: str = "") -> "DrawStorage":
        """Create DrawStorage from a Roster object."""
        games = []
        game_id = 0
        
        for weekly_draw in roster.weeks:
            for game in weekly_draw.games:
                stored_game = StoredGame(
                    game_id=f"G{game_id:05d}",
                    team1=game.team1,
                    team2=game.team2,
                    grade=game.grade.name,
                    week=weekly_draw.week,
                    round_no=weekly_draw.round_no,
                    date=game.timeslot.date,
                    day=game.timeslot.day,
                    time=game.timeslot.time,
                    day_slot=game.timeslot.day_slot,
                    field_name=game.field.name,
                    field_location=game.field.location
                )
                games.append(stored_game)
                game_id += 1
        
        weeks = set(g.week for g in games)
        
        return cls(
            description=description,
            num_weeks=len(weeks),
            num_games=len(games),
            games=games
        )
    
    @classmethod
    def load(cls, path: str) -> "DrawStorage":
        """Load from JSON file."""
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return cls(**data)
    
    def save(self, path: str) -> None:
        """Save to JSON file."""
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(self.model_dump(), f, indent=2)
        print(f"Draw saved to {path}")
    
    def to_X_dict(self) -> Dict[Tuple, int]:
        """Convert back to X solution dict format."""
        return {game.to_key(): 1 for game in self.games}
    
    def get_game_by_id(self, game_id: str) -> Optional[StoredGame]:
        """Get a game by its ID."""
        for game in self.games:
            if game.game_id == game_id:
                return game
        return None
    
    def get_games_by_team(self, team_name: str) -> List[StoredGame]:
        """Get all games for a team."""
        return [g for g in self.games if team_name in (g.team1, g.team2)]
    
    def get_games_by_club(self, club_name: str) -> List[StoredGame]:
        """Get all games involving a club's teams."""
        return [g for g in self.games 
                if club_name in g.team1 or club_name in g.team2]
    
    def get_games_by_week(self, week: int) -> List[StoredGame]:
        """Get all games for a specific week."""
        return [g for g in self.games if g.week == week]
    
    def get_games_by_grade(self, grade: str) -> List[StoredGame]:
        """Get all games for a specific grade."""
        return [g for g in self.games if g.grade == grade]
    
    def filter_games(
        self,
        team: Optional[str] = None,
        grade: Optional[str] = None,
        week: Optional[int] = None,
        field_location: Optional[str] = None
    ) -> List[StoredGame]:
        """Filter games by multiple criteria."""
        result = self.games
        if team:
            result = [g for g in result if team in (g.team1, g.team2)]
        if grade:
            result = [g for g in result if g.grade == grade]
        if week:
            result = [g for g in result if g.week == week]
        if field_location:
            result = [g for g in result if g.field_location == field_location]
        return result
    
    # ============== Partial Draw Operations ==============
    
    def get_locked_games(self, locked_weeks) -> List[StoredGame]:
        """Get games in weeks that should be locked.
        
        Args:
            locked_weeks: A set/list of week numbers to lock, or an int (locks weeks 1..N).
        """
        if isinstance(locked_weeks, int):
            locked_weeks = set(range(1, locked_weeks + 1))
        locked_weeks = set(locked_weeks)
        return [g for g in self.games if g.week in locked_weeks]
    
    def get_remaining_games(self, locked_weeks) -> List[StoredGame]:
        """Get games in weeks that are NOT locked.
        
        Args:
            locked_weeks: A set/list of week numbers to lock, or an int (locks weeks 1..N).
        """
        if isinstance(locked_weeks, int):
            locked_weeks = set(range(1, locked_weeks + 1))
        locked_weeks = set(locked_weeks)
        return [g for g in self.games if g.week not in locked_weeks]
    
    def lock_and_split(self, locked_weeks) -> Tuple["DrawStorage", "DrawStorage"]:
        """
        Split draw into locked and unlocked portions.
        
        Args:
            locked_weeks: A set/list of week numbers to lock, or an int (locks weeks 1..N).
            
        Returns:
            Tuple of (locked_draw, remaining_draw)
        """
        if isinstance(locked_weeks, int):
            locked_weeks = set(range(1, locked_weeks + 1))
        locked_weeks = set(locked_weeks)
        
        locked_games = self.get_locked_games(locked_weeks)
        remaining_games = self.get_remaining_games(locked_weeks)
        
        weeks_label = ','.join(str(w) for w in sorted(locked_weeks))
        locked_draw = DrawStorage(
            description=f"{self.description} (Locked weeks {weeks_label})",
            num_weeks=len(locked_weeks),
            num_games=len(locked_games),
            games=locked_games,
            metadata={**self.metadata, 'locked_weeks': sorted(locked_weeks)}
        )
        
        remaining_week_nums = set(g.week for g in remaining_games)
        remaining_draw = DrawStorage(
            description=f"{self.description} (Remaining weeks)",
            num_weeks=len(remaining_week_nums),
            num_games=len(remaining_games),
            games=remaining_games,
            metadata={**self.metadata, 'unlocked_weeks': sorted(remaining_week_nums)}
        )
        
        return locked_draw, remaining_draw
    
    @classmethod
    def load_and_lock(
        cls,
        path: str,
        locked_weeks
    ) -> Tuple["DrawStorage", List[Tuple]]:
        """
        Load a draw and prepare locked games for solver constraint injection.
        
        Args:
            path: Path to JSON draw file
            locked_weeks: A set/list of week numbers to lock, or an int (locks weeks 1..N).
            
        Returns:
            Tuple of (locked_draw, locked_game_keys)
            - locked_draw: DrawStorage containing only locked games
            - locked_game_keys: List of 11-tuple keys for model.Add(X[key] == 1)
        """
        if isinstance(locked_weeks, int):
            locked_weeks = set(range(1, locked_weeks + 1))
        locked_weeks = set(locked_weeks)
        
        full_draw = cls.load(path)
        locked_games = full_draw.get_locked_games(locked_weeks)
        
        weeks_label = ','.join(str(w) for w in sorted(locked_weeks))
        locked_draw = DrawStorage(
            description=f"Locked from {full_draw.description} (weeks {weeks_label})",
            num_weeks=len(locked_weeks),
            num_games=len(locked_games),
            games=locked_games,
            metadata={'source_file': path, 'locked_weeks': sorted(locked_weeks)}
        )
        
        locked_keys = [game.to_key() for game in locked_games]
        
        print(f"Loaded {len(locked_games)} locked games from weeks {weeks_label}")
        return locked_draw, locked_keys

    @classmethod
    def from_excel(
        cls,
        path: str,
        data: Dict,
        description: str = ""
    ) -> "DrawStorage":
        """
        Load a draw from Excel file format (weekly sheets).
        
        Args:
            path: Path to Excel file
            data: Data dict containing teams, grades, fields
            description: Optional description
            
        Returns:
            DrawStorage object
        """
        import pandas as pd
        
        games = []
        game_id = 0
        
        # Build field lookup
        fields = {f.name: f for f in data.get('fields', [])}
        grades = {g.name: g for g in data.get('grades', [])}
        
        with pd.ExcelFile(path) as xlsx:
            for sheet_name in xlsx.sheet_names:
                if not sheet_name.startswith('Week '):
                    continue
                    
                try:
                    week_num = int(sheet_name.replace('Week ', ''))
                except ValueError:
                    continue
                
                df = pd.read_excel(xlsx, sheet_name=sheet_name)
                
                for _, row in df.iterrows():
                    # Skip empty rows or bye markers
                    if pd.isna(row.get('Team 1')) or str(row.get('Team 1')).strip().lower() == 'byes and no games':
                        break
                    
                    try:
                        game = StoredGame(
                            game_id=f"G{game_id:05d}",
                            team1=str(row['Team 1']).strip(),
                            team2=str(row['Team 2']).strip(),
                            grade=str(row['Grade']).strip(),
                            week=week_num,
                            round_no=int(row.get('Round', week_num)),
                            date=str(row['Date']).strip(),
                            day=str(row['Day']).strip(),
                            time=str(row['Time']).strip(),
                            day_slot=int(row.get('Day Slot', 1)),
                            field_name=str(row['Field Name']).strip(),
                            field_location=str(row['Field Location']).strip()
                        )
                        games.append(game)
                        game_id += 1
                    except Exception as e:
                        print(f"Warning: Skipping row in {sheet_name}: {e}")
                        continue
        
        weeks = set(g.week for g in games)
        
        return cls(
            description=description or f"Imported from {path}",
            num_weeks=len(weeks),
            num_games=len(games),
            games=games
        )
    
    def merge_with(self, other: "DrawStorage") -> "DrawStorage":
        """
        Merge another draw into this one.
        
        Useful for combining locked games with newly solved games.
        
        Args:
            other: Another DrawStorage to merge
            
        Returns:
            New DrawStorage with combined games
        """
        all_games = self.games + other.games
        
        # Re-number game IDs
        for i, game in enumerate(all_games):
            game_dict = game.model_dump()
            game_dict['game_id'] = f"G{i:05d}"
            all_games[i] = StoredGame(**game_dict)
        
        weeks = set(g.week for g in all_games)
        
        return DrawStorage(
            description=f"Merged: {self.description} + {other.description}",
            num_weeks=len(weeks),
            num_games=len(all_games),
            games=all_games,
            metadata={
                'merge_sources': [self.description, other.description],
                'merged_at': datetime.now().isoformat()
            }
        )

    # ============== Excel Export ==============

    def export_schedule_xlsx(
        self,
        filename: str,
        weeks: Optional[List[int]] = None,
        sheet_title: Optional[str] = None,
        weekend_notes: Optional[Dict[int, List[str]]] = None,
    ) -> None:
        """Export draw to a nicely formatted Excel schedule.

        Games are grouped by week → date → field with colour-coded headers,
        field sub-headers, borders, and bye listings.  Works directly from
        DrawStorage (no Roster / ``data`` dict needed).

        When ``weekend_notes`` is supplied (a ``{week: [note_line, ...]}`` dict
        as returned by ``analytics.notes.build_weekend_notes``), a Notes column
        is written in column N.  Columns L and M are left blank as spacers.
        When ``weekend_notes`` is ``None`` (the default) the output is identical
        to the pre-notes behaviour in columns A–K and nothing is written to L–N.

        Args:
            filename: Output .xlsx path.
            weeks: Optional list of week numbers to include.  ``None`` = all.
            sheet_title: Worksheet title (default auto-generated).
            weekend_notes: Optional mapping of week number → list of formatted
                note strings (e.g. ``"Field: Masters SC weekend"``).  When
                supplied, notes are stacked from the week's first game row
                downward in column N.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        # --- filter games ---------------------------------------------------
        games = list(self.games)
        if weeks is not None:
            week_set = set(weeks)
            games = [g for g in games if g.week in week_set]

        by_week: Dict[int, list] = defaultdict(list)
        for g in games:
            by_week[g.week].append(g)

        # all teams (for byes) scoped to selected weeks' grades
        all_teams_by_grade: Dict[str, set] = defaultdict(set)
        source = self.games  # full draw for bye calculation
        for g in source:
            all_teams_by_grade[g.grade].add(g.team1)
            all_teams_by_grade[g.grade].add(g.team2)

        # --- styles ----------------------------------------------------------
        WEEK_COLOURS = [
            'D6E4F0',  # light blue
            'E2EFDA',  # light green
            'FCE4D6',  # light peach
            'EDEDED',  # light grey
            'FFF2CC',  # light yellow
            'D9E2F3',  # periwinkle
            'E2D9F3',  # light purple
            'D6F0E4',  # light teal
        ]
        field_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        field_font = Font(bold=True, size=10, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(bold=True, size=10, color='FFFFFF')
        week_font = Font(bold=True, size=13)
        bye_font = Font(bold=True, size=10, color='CC0000')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        HEADERS = ['Week', 'Round', 'Date', 'Day', 'Time', 'Slot',
                    'Team 1', 'Team 2', 'Grade', 'Field', 'Venue']
        COL_WIDTHS = {
            'A': 8, 'B': 8, 'C': 14, 'D': 10, 'E': 8, 'F': 6,
            'G': 28, 'H': 28, 'I': 8, 'J': 8, 'K': 38,
            'L': 3, 'M': 3, 'N': 60,
        }
        FIELD_ORDER = {'EF': 0, 'WF': 1, 'SF': 2}
        NUM_COLS = len(HEADERS)
        # Notes column index: two blank spacer columns (L=NUM_COLS+1, M=NUM_COLS+2)
        # then Notes in N=NUM_COLS+3.  Using the constant means it tracks HEADERS.
        notes_col = NUM_COLS + 3  # = 14 when NUM_COLS=11

        # --- build workbook --------------------------------------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        if sheet_title:
            ws.title = sheet_title
        elif weeks:
            ws.title = f"Weeks {min(weeks)}-{max(weeks)}"
        else:
            ws.title = "Schedule"

        row = 1
        for idx, week in enumerate(sorted(by_week.keys())):
            week_games = by_week[week]
            week_colour = WEEK_COLOURS[idx % len(WEEK_COLOURS)]
            week_bg = PatternFill(start_color=week_colour,
                                  end_color=week_colour, fill_type='solid')

            # Week title row
            title_row = row
            c = ws.cell(row=row, column=1, value=f'Week {week}')
            c.font = week_font
            for ci in range(1, NUM_COLS + 1):
                ws.cell(row=row, column=ci).fill = week_bg
            # When notes active: extend the week background colour to column N
            # so the title band visually spans to the notes column.
            if weekend_notes is not None:
                ws.cell(row=title_row, column=notes_col).fill = week_bg
            row += 1

            # Column headers
            header_row = row
            for ci, h in enumerate(HEADERS, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.border = thin_border
            # Write Notes header in column N (separate from HEADERS so NUM_COLS
            # is not changed and the week-title fill range A–K stays correct).
            if weekend_notes is not None:
                c = ws.cell(row=header_row, column=notes_col, value="Notes")
                c.font = header_font
                c.fill = header_fill
                c.border = thin_border
            row += 1

            # Track the first game data row within this week (for note stacking).
            first_game_row: Optional[int] = None

            # Group by date → field
            by_date: Dict[str, list] = defaultdict(list)
            for g in week_games:
                by_date[g.date].append(g)

            for date in sorted(by_date.keys()):
                date_games = by_date[date]
                by_field: Dict[tuple, list] = defaultdict(list)
                for g in date_games:
                    by_field[(g.field_name, g.field_location)].append(g)

                sorted_fields = sorted(
                    by_field.keys(),
                    key=lambda f: (
                        0 if 'Newcastle' in f[1] else 1,
                        FIELD_ORDER.get(f[0], 99),
                        f[0],
                    ),
                )

                for field_name, field_loc in sorted_fields:
                    field_games = sorted(by_field[(field_name, field_loc)],
                                         key=lambda g: g.time)
                    # Field sub-header
                    c = ws.cell(row=row, column=1,
                                value=f'{field_name} - {field_loc}')
                    c.font = field_font
                    for ci in range(1, NUM_COLS + 1):
                        ws.cell(row=row, column=ci).fill = field_fill
                    row += 1

                    # Game rows
                    for g in field_games:
                        # Capture the first game data row for note placement.
                        if first_game_row is None:
                            first_game_row = row
                        vals = [g.week, g.round_no, g.date, g.day, g.time,
                                g.day_slot, g.team1, g.team2, g.grade,
                                g.field_name, g.field_location]
                        for ci, v in enumerate(vals, 1):
                            c = ws.cell(row=row, column=ci, value=v)
                            c.border = thin_border
                            c.fill = week_bg
                        row += 1

            # Write stacked notes for this week in column N (A–K are never
            # touched here; notes live in their own column).
            if weekend_notes is not None and first_game_row is not None:
                for i, note_text in enumerate(weekend_notes.get(week, [])):
                    note_row = first_game_row + i
                    c = ws.cell(row=note_row, column=notes_col, value=note_text)
                    c.font = Font(italic=False)
                    c.border = thin_border
                    c.fill = week_bg

            # Byes
            playing: Dict[str, set] = defaultdict(set)
            for g in week_games:
                playing[g.grade].add(g.team1)
                playing[g.grade].add(g.team2)
            bye_teams = []
            for grade in ['PHL', '2nd', '3rd', '4th', '5th', '6th']:
                for team in sorted(
                    all_teams_by_grade.get(grade, set())
                    - playing.get(grade, set())
                ):
                    bye_teams.append(team)

            row += 1  # blank row
            if bye_teams:
                c = ws.cell(row=row, column=1, value='BYES')
                c.font = bye_font
                row += 1
                for team in bye_teams:
                    ws.cell(row=row, column=1, value=team).font = Font(
                        italic=True)
                    row += 1

            row += 1  # gap between weeks

        # Column widths
        for col_letter, width in COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(filename)
        print(f"Schedule exported to {filename}")


# ============== Draw Analytics ==============

class DrawAnalytics:
    """Comprehensive analytics generator for draw schedules."""
    
    def __init__(self, draw: DrawStorage, data: Dict):
        """
        Initialize analytics.
        
        Args:
            draw: DrawStorage object containing the draw
            data: Data dictionary containing teams, grades, clubs, etc.
        """
        self.draw = draw
        self.data = data
        self.teams: List[Team] = data.get('teams', [])
        self.grades: List[Grade] = data.get('grades', [])
        self.clubs: List[Club] = data.get('clubs', [])
        
        # Build lookup tables
        self._team_to_club = {t.name: t.club.name for t in self.teams}
        self._team_to_grade = {t.name: t.grade for t in self.teams}
        self._club_teams = defaultdict(list)
        for t in self.teams:
            self._club_teams[t.club.name].append(t.name)
    
    def get_club_for_team(self, team_name: str) -> str:
        """Get club name for a team."""
        return self._team_to_club.get(team_name, "Unknown")
    
    # ---------- Games Played Cross Tab (Grade x Team) ----------
    
    def games_played_by_team_grade(self) -> pd.DataFrame:
        """
        Create cross-tab of games played per team per grade.
        Returns DataFrame with grades as columns, teams as rows.
        """
        counts = defaultdict(lambda: defaultdict(int))
        
        for game in self.draw.games:
            counts[game.team1][game.grade] += 1
            counts[game.team2][game.grade] += 1
        
        # Convert to DataFrame
        grade_names = [g.name for g in self.grades]
        df_data = []
        
        for team in self.teams:
            row = {'Team': team.name, 'Club': team.club.name, 'Grade': team.grade}
            for grade in grade_names:
                row[grade] = counts[team.name].get(grade, 0)
            row['Total'] = sum(row.get(g, 0) for g in grade_names)
            df_data.append(row)
        
        df = pd.DataFrame(df_data)
        return df.sort_values(['Club', 'Grade', 'Team'])
    
    # ---------- Team vs Team Matchups (Grade x Team x Team) ----------
    
    def team_matchups_crosstab(self, grade: Optional[str] = None) -> Dict[str, pd.DataFrame]:
        """
        Create matchup matrix for each grade.
        Returns dict of DataFrames, one per grade.
        """
        # Group games by grade
        games_by_grade = defaultdict(list)
        for game in self.draw.games:
            games_by_grade[game.grade].append(game)
        
        result = {}
        grades_to_process = [grade] if grade else [g.name for g in self.grades]
        
        for grade_name in grades_to_process:
            if grade_name not in games_by_grade:
                continue
            
            grade_games = games_by_grade[grade_name]
            
            # Get teams in this grade
            grade_teams = sorted([t.name for t in self.teams if t.grade == grade_name])
            
            # Build matchup matrix
            matchups = defaultdict(lambda: defaultdict(int))
            for game in grade_games:
                matchups[game.team1][game.team2] += 1
                matchups[game.team2][game.team1] += 1
            
            # Convert to DataFrame
            df_data = []
            for team1 in grade_teams:
                row = {'Team': team1}
                for team2 in grade_teams:
                    if team1 == team2:
                        row[team2] = '-'
                    else:
                        row[team2] = matchups[team1].get(team2, 0)
                df_data.append(row)
            
            df = pd.DataFrame(df_data)
            df = df.set_index('Team')
            result[grade_name] = df
        
        return result
    
    # ---------- Home vs Away Analysis ----------
    
    def home_away_analysis(self) -> pd.DataFrame:
        """
        Create home vs away analysis for each team.
        Shows games played at home field vs away.
        """
        team_stats = defaultdict(lambda: {
            'home': 0, 'away': 0, 'neutral': 0, 
            'vs_maitland_home': 0, 'at_maitland': 0
        })
        
        for game in self.draw.games:
            t1_club = self.get_club_for_team(game.team1)
            t2_club = self.get_club_for_team(game.team2)
            
            # Determine home/away based on field location
            location = game.field_location
            
            # Team 1 analysis
            if self._is_home_location(t1_club, location):
                team_stats[game.team1]['home'] += 1
                team_stats[game.team2]['away'] += 1
            elif self._is_home_location(t2_club, location):
                team_stats[game.team2]['home'] += 1
                team_stats[game.team1]['away'] += 1
            else:
                # Neutral (Broadmeadow) - both away from home
                team_stats[game.team1]['neutral'] += 1
                team_stats[game.team2]['neutral'] += 1
            
            # Track Maitland specific stats
            if 'Maitland' in t1_club or 'Maitland' in t2_club:
                if location == 'Maitland Park':
                    for team in [game.team1, game.team2]:
                        if 'Maitland' not in team:
                            team_stats[team]['at_maitland'] += 1
        
        # Convert to DataFrame
        df_data = []
        for team in self.teams:
            stats = team_stats[team.name]
            total = stats['home'] + stats['away'] + stats['neutral']
            df_data.append({
                'Team': team.name,
                'Club': team.club.name,
                'Grade': team.grade,
                'Home': stats['home'],
                'Away': stats['away'],
                'Neutral': stats['neutral'],
                'Total': total,
                'Home%': f"{100*stats['home']/total:.1f}%" if total > 0 else "0%",
                'At Maitland': stats['at_maitland']
            })
        
        df = pd.DataFrame(df_data)
        return df.sort_values(['Club', 'Grade', 'Team'])
    
    def _is_home_location(self, club: str, location: str) -> bool:
        """Check if a location is a club's home field."""
        home_mappings = {
            'Maitland': 'Maitland Park',
            'Gosford': 'Central Coast Hockey Park',
            'University': 'University No.1',
            # Add other club-to-field mappings as needed
        }
        return home_mappings.get(club, '') == location
    
    # ---------- Club Season Schedule ----------
    
    def club_season_schedule(self, club_name: str) -> pd.DataFrame:
        """
        Create a complete season schedule for a club.
        Shows all teams, all weeks, with game details.
        """
        club_teams = [t.name for t in self.teams if t.club.name == club_name]
        club_games = [g for g in self.draw.games 
                     if g.team1 in club_teams or g.team2 in club_teams]
        
        df_data = []
        for game in sorted(club_games, key=lambda g: (g.week, g.day_slot)):
            # Determine if home/away for this club
            is_home = game.team1 in club_teams
            club_team = game.team1 if is_home else game.team2
            opponent = game.team2 if is_home else game.team1
            
            df_data.append({
                'Week': game.week,
                'Round': game.round_no,
                'Date': game.date,
                'Day': game.day,
                'Time': game.time,
                'Team': club_team,
                'Grade': game.grade,
                'H/A': 'Home' if is_home else 'Away',
                'Opponent': opponent,
                'Field': game.field_name,
                'Location': game.field_location
            })
        
        return pd.DataFrame(df_data)
    
    # ---------- Grade Summary ----------
    
    def grade_summary(self) -> pd.DataFrame:
        """Summary statistics per grade."""
        grade_stats = defaultdict(lambda: {
            'num_games': 0,
            'teams': set(),
            'matchups': set()
        })
        
        for game in self.draw.games:
            grade_stats[game.grade]['num_games'] += 1
            grade_stats[game.grade]['teams'].add(game.team1)
            grade_stats[game.grade]['teams'].add(game.team2)
            pair = tuple(sorted([game.team1, game.team2]))
            grade_stats[game.grade]['matchups'].add(pair)
        
        df_data = []
        for grade in self.grades:
            stats = grade_stats.get(grade.name, {'num_games': 0, 'teams': set(), 'matchups': set()})
            df_data.append({
                'Grade': grade.name,
                'Num Teams': len(stats['teams']),
                'Total Games': stats['num_games'],
                'Unique Matchups': len(stats['matchups'])
            })
        
        return pd.DataFrame(df_data)
    
    # ---------- Weekly Field Usage ----------
    
    def weekly_field_usage(self) -> pd.DataFrame:
        """Analyze field usage per week."""
        field_usage = defaultdict(lambda: defaultdict(int))
        
        for game in self.draw.games:
            key = (game.field_location, game.field_name)
            field_usage[game.week][key] += 1
        
        df_data = []
        weeks = sorted(set(g.week for g in self.draw.games))
        
        for week in weeks:
            week_usage = field_usage[week]
            for (location, field), count in week_usage.items():
                df_data.append({
                    'Week': week,
                    'Location': location,
                    'Field': field,
                    'Games': count
                })
        
        return pd.DataFrame(df_data).sort_values(['Week', 'Location', 'Field'])
    
    # ---------- Maitland/Away Team Analysis ----------
    
    def away_team_balance(self) -> pd.DataFrame:
        """
        Analyze home/away balance for Maitland and Gosford teams.
        Critical for ensuring 50/50 balance constraint is met.
        """
        away_teams = ['Maitland', 'Gosford']
        
        df_data = []
        for away_prefix in away_teams:
            # Get teams from this club
            club_teams = [t.name for t in self.teams if away_prefix in t.name]
            
            for team_name in club_teams:
                team_obj = next((t for t in self.teams if t.name == team_name), None)
                if not team_obj:
                    continue
                
                home_games = 0
                away_games = 0
                
                for game in self.draw.games:
                    if team_name not in (game.team1, game.team2):
                        continue
                    
                    # Get opponent
                    opponent = game.team2 if game.team1 == team_name else game.team1
                    
                    # Skip intra-club games
                    if away_prefix in opponent:
                        continue
                    
                    # Determine home field
                    home_location = 'Maitland Park' if away_prefix == 'Maitland' else 'Central Coast Hockey Park'
                    
                    if game.field_location == home_location:
                        home_games += 1
                    else:
                        away_games += 1
                
                total = home_games + away_games
                df_data.append({
                    'Team': team_name,
                    'Club': away_prefix,
                    'Grade': team_obj.grade,
                    'Home Games': home_games,
                    'Away Games': away_games,
                    'Total': total,
                    'Home%': f"{100*home_games/total:.1f}%" if total > 0 else "N/A",
                    'Balanced': 'Yes' if abs(home_games - away_games) <= 1 else 'NO'
                })
        
        return pd.DataFrame(df_data)
    
    # ---------- Constraint Compliance Summary ----------
    
    def constraint_compliance_summary(self) -> pd.DataFrame:
        """Quick compliance check for key constraints."""
        checks = []
        
        # Check 1: Teams per week (should be exactly 1)
        team_games_per_week = defaultdict(int)
        for game in self.draw.games:
            team_games_per_week[(game.week, game.team1)] += 1
            team_games_per_week[(game.week, game.team2)] += 1
        
        double_bookings = [k for k, v in team_games_per_week.items() if v > 1]
        checks.append({
            'Constraint': 'No Double Booking (Teams)',
            'Status': 'PASS' if not double_bookings else 'FAIL',
            'Issues': len(double_bookings)
        })
        
        # Check 2: Fields per slot
        field_games_per_slot = defaultdict(int)
        for game in self.draw.games:
            field_games_per_slot[(game.week, game.day_slot, game.field_name)] += 1
        
        field_conflicts = [k for k, v in field_games_per_slot.items() if v > 1]
        checks.append({
            'Constraint': 'No Double Booking (Fields)',
            'Status': 'PASS' if not field_conflicts else 'FAIL',
            'Issues': len(field_conflicts)
        })
        
        # Check 3: Maitland back-to-back
        maitland_home_weeks = set()
        for game in self.draw.games:
            if ('Maitland' in game.team1 or 'Maitland' in game.team2) and game.field_location == 'Maitland Park':
                maitland_home_weeks.add(game.week)
        
        consecutive = 0
        sorted_weeks = sorted(maitland_home_weeks)
        for i in range(1, len(sorted_weeks)):
            if sorted_weeks[i] == sorted_weeks[i-1] + 1:
                consecutive += 1
        
        checks.append({
            'Constraint': 'No Back-to-Back Maitland Home',
            'Status': 'PASS' if consecutive == 0 else 'FAIL',
            'Issues': consecutive
        })
        
        return pd.DataFrame(checks)
    
    # ---------- Export to Excel ----------
    
    def export_analytics_to_excel(self, filename: str = "draw_analytics.xlsx") -> None:
        """Export all analytics to a comprehensive Excel file."""
        with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
            # 1. Summary sheet
            summary = self.grade_summary()
            summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # 2. Compliance check
            compliance = self.constraint_compliance_summary()
            compliance.to_excel(writer, sheet_name='Compliance Check', index=False)
            
            # 3. Games per team/grade
            games_by_team = self.games_played_by_team_grade()
            games_by_team.to_excel(writer, sheet_name='Games Per Team', index=False)
            
            # 4. Home/Away analysis
            home_away = self.home_away_analysis()
            home_away.to_excel(writer, sheet_name='Home-Away Analysis', index=False)
            
            # 5. Away team balance (Maitland/Gosford)
            away_balance = self.away_team_balance()
            away_balance.to_excel(writer, sheet_name='Away Team Balance', index=False)
            
            # 6. Matchup matrices per grade
            matchups = self.team_matchups_crosstab()
            row_offset = 0
            for grade, df in matchups.items():
                df.to_excel(writer, sheet_name='Matchup Matrix', startrow=row_offset)
                row_offset += len(df) + 3  # Gap between grades
            
            # 7. Weekly field usage
            field_usage = self.weekly_field_usage()
            field_usage.to_excel(writer, sheet_name='Field Usage', index=False)
            
            # 8. Club schedules (one sheet per club)
            for club in self.clubs:
                schedule = self.club_season_schedule(club.name)
                if not schedule.empty:
                    sheet_name = f"Club-{club.name[:20]}"  # Excel sheet name limit
                    schedule.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 9. Full game list
            full_games = pd.DataFrame([g.model_dump() for g in self.draw.games])
            if not full_games.empty:
                full_games = full_games.sort_values(['week', 'day_slot', 'grade'])
                full_games.to_excel(writer, sheet_name='All Games', index=False)
        
        print(f"Analytics exported to {filename}")


# ============== Utility Functions ==============

def extract_locked_pairings(
    draw: "DrawStorage",
    *,
    freeze_grades,
    freeze_weeks,
    exclude_weeks=frozenset(),
) -> List[Dict[str, Any]]:
    """Return a LOCKED_PAIRINGS list for use with the regen path (spec-026 DoD #1).

    Each entry is a dict with keys ``teams``, ``grade``, ``date``, and
    ``description`` — the same shape as a hand-authored LOCKED_PAIRINGS entry.
    The slot/time/field keys are intentionally omitted so the solver is free to
    re-time the game within the pinned date.

    A game is included iff ALL of the following hold:
      1. ``g.grade in freeze_grades``  (grade axis: frozen grade)
      2. ``g.week  in freeze_weeks``   (week axis: frozen week)
      3. ``g.week  not in exclude_weeks`` (played weeks, already hard-locked, must
         not be double-pinned via LOCKED_PAIRINGS)

    Empty-set semantics: an empty ``freeze_grades`` or ``freeze_weeks`` means
    nothing matches along that axis, so the function returns [].

    Bye/placeholder games — where ``team1`` or ``team2`` is falsy — are silently
    skipped.  Real DrawStorage objects don't store bye entries as StoredGame
    records (byes are computed dynamically in export), but this guard protects
    against manually-constructed DrawStorage objects with a missing opponent.

    Args:
        draw: The source DrawStorage to extract pins from.
        freeze_grades: Iterable of grade strings to freeze (e.g. ``{'PHL', '2nd'}``).
        freeze_weeks: Iterable of week ints to freeze (e.g. ``{1, 2, 3}``).
        exclude_weeks: Iterable of week ints to EXCLUDE from pinning even if they
            would otherwise match.  Typically these are the hard-locked
            already-played weeks passed via ``--lock-weeks``.  Defaults to the
            empty frozenset (nothing excluded).

    Returns:
        A list of dicts shaped ``{'teams': [t1, t2], 'grade': str, 'date': str,
        'description': str}``, one per frozen game.
    """
    freeze_grades_set = set(freeze_grades)
    freeze_weeks_set = set(freeze_weeks)
    exclude_weeks_set = set(exclude_weeks)

    pins: List[Dict[str, Any]] = []
    for g in draw.games:
        # Bye / placeholder guard
        if not g.team1 or not g.team2:
            continue
        # Grade and week membership tests
        if g.grade not in freeze_grades_set:
            continue
        if g.week not in freeze_weeks_set:
            continue
        # Exclude hard-locked played weeks to avoid double-pinning
        if g.week in exclude_weeks_set:
            continue
        pins.append({
            'teams': [g.team1, g.team2],
            'grade': g.grade,
            'date': g.date,
            'description': f"Regen pin: {g.team1} vs {g.team2} ({g.grade}) {g.date}",
        })
    return pins


def count_games_changed(source: "DrawStorage", result: "DrawStorage") -> int:
    """Count games whose time/slot/field differ between two draws (spec-026 DoD #6).

    Two draws are matched game-by-game on the *pairing identity*
    ``(team1, team2, grade, week)``.  For each pairing present in BOTH draws,
    the game is counted as "changed" if any of its schedule fields differ:
    ``time``, ``day_slot``, ``field_name``, or ``field_location``.

    Pairings present in only one draw (added or removed) are NOT counted here —
    this helper measures the re-timing footprint of a regen, not pairing churn.
    A regen that freezes a grade keeps its pairings on the same date but may
    re-time them; this helper counts exactly those re-timed games.

    Args:
        source: The source (pre-regen) draw.
        result: The result (post-regen) draw.

    Returns:
        The number of common pairings whose time/slot/field changed.
    """
    def _ident(g: "StoredGame"):
        return (g.team1, g.team2, g.grade, g.week)

    def _sched(g: "StoredGame"):
        return (g.time, g.day_slot, g.field_name, g.field_location)

    src_by_ident = {_ident(g): _sched(g) for g in source.games}
    changed = 0
    for g in result.games:
        ident = _ident(g)
        if ident in src_by_ident and src_by_ident[ident] != _sched(g):
            changed += 1
    return changed


def create_draw_from_solution(X_solution: Dict, description: str = "") -> DrawStorage:
    """Convenience function to create DrawStorage from X solution."""
    return DrawStorage.from_X_solution(X_solution, description)


def load_draw(path: str) -> DrawStorage:
    """Load a draw from JSON file."""
    return DrawStorage.load(path)


def analyze_draw(draw: DrawStorage, data: Dict) -> DrawAnalytics:
    """Create analytics for a draw."""
    return DrawAnalytics(draw, data)


# ============== Rev Format Export ==============

def export_draw_to_revformat(
    draw: DrawStorage,
    data: Dict,
    output_path: str = "draws/2026/revformat/schedule_revformat.csv",
    week_limit: Optional[int] = None
) -> None:
    """
    Export draw to rev format CSV for integration with external systems.
    
    Args:
        draw: DrawStorage containing the schedule
        data: Data dict containing teams for bye information
        output_path: Output CSV file path
        week_limit: Only export up to this week (optional)
    """
    from datetime import datetime as dt
    
    # Mappings for external system format
    grade_map = {
        'PHL': 'HCPHL', 
        '2nd': '2ND GRADE', 
        '3rd': '3RD GRADE', 
        '4th': '4TH GRADE', 
        '5th': '5TH GRADE', 
        '6th': '6TH GRADE'
    }
    field_map = {'EF': 'EF', 'WF': 'WF', 'SF': 'SF', 'Maitland Main Field': '', 'Wyong Main Field': ''}
    
    # Revo full club name mapping: short club name -> full registered club name
    revo_club_names = {
        'Crusaders': 'Crusaders Hockey Club (Mens & Womens)',
        'Gosford': 'Gosford City Hockey Club (Mens)',
        'Maitland': 'Maitland District Hockey Club',
        'Tigers': 'Newcastle Tigers Hockey Club',
        'Souths': 'South Newcastle Mens Hockey Club',
        'Norths': 'Phoenix North Hockey Club',
        'Port Stephens': 'Port Stephens Hornets Hockey Club Inc (Mens)',
        'University': 'University of Newcastle Men\'s Hockey Club',
        'Wests': 'West Hockey Club',
        'Colts': 'Colts Mens Hockey Club',
        'Cardiff': 'Cardiff Hockey Club',
    }
    
    def to_revo_team_name(team_with_grade: str, grade: str) -> str:
        """Convert internal team name to Revo format.
        
        e.g. 'Crusaders 4th' -> 'Crusaders Hockey Club (Mens & Womens) Crusaders'
             'Tigers Black 6th' -> 'Newcastle Tigers Hockey Club Tigers Black'
             'University Redhogs 4th' -> "University of Newcastle Men's Hockey Club University Redhogs"
        """
        # Strip grade suffix
        suffix = f" {grade}"
        short_name = team_with_grade[:-len(suffix)] if team_with_grade.endswith(suffix) else team_with_grade
        
        # Find the matching club by checking which club name the short_name starts with
        # Sort by length descending so "Port Stephens" matches before a hypothetical "Port"
        for club_short in sorted(revo_club_names.keys(), key=len, reverse=True):
            if short_name == club_short or short_name.startswith(club_short + ' '):
                return f"{revo_club_names[club_short]} {short_name}"
        
        # Fallback: use short name as-is with generic prefix
        return f"Newcastle Hockey Association {short_name}"
    
    all_rows = []

    # Get all teams for bye tracking
    teams_by_grade = defaultdict(set)
    for team in data.get('teams', []):
        teams_by_grade[team.grade].add(team.name)

    # Group games by week for bye calculation
    games_by_week = defaultdict(list)
    for game in draw.games:
        if week_limit and game.week > week_limit:
            continue
        games_by_week[game.week].append(game)

    default_venue = 'Newcastle International Hockey Centre'

    for week, games in sorted(games_by_week.items()):
        # Track teams playing this week by grade
        teams_playing = defaultdict(set)

        for game in games:
            teams_playing[game.grade].add(game.team1)
            teams_playing[game.grade].add(game.team2)

            team1_revo = to_revo_team_name(game.team1, game.grade)
            team2_revo = to_revo_team_name(game.team2, game.grade)

            all_rows.append({
                'Round': game.round_no,
                'Round type': '',
                'Grade': grade_map.get(game.grade, game.grade),
                'Game date': game.date,
                'Venue': game.field_location,
                'Subvenue': field_map.get(game.field_name, game.field_name),
                'Time': game.time,
                'Pool': '',
                'Team 1 ': team1_revo,
                'Team 1 score': '',
                'Team 2': team2_revo,
                'Team 2 score': '',
                'Duty team': '',
                'Label': '',
            })

        # Add bye entries
        for grade, playing in teams_playing.items():
            all_teams = teams_by_grade.get(grade, set())
            bye_teams = all_teams - playing
            for team in sorted(bye_teams):
                team_revo = to_revo_team_name(team, grade)
                ref_game = games[0] if games else None
                if ref_game:
                    all_rows.append({
                        'Round': ref_game.round_no,
                        'Round type': '',
                        'Grade': grade_map.get(grade, grade),
                        'Game date': ref_game.date,
                        'Venue': default_venue,
                        'Subvenue': '',
                        'Time': '',
                        'Pool': '',
                        'Team 1 ': team_revo,
                        'Team 1 score': '',
                        'Team 2': 'BYE',
                        'Team 2 score': '',
                        'Duty team': '',
                        'Label': '',
                    })

    # Create DataFrame and sort
    columns = ['Round', 'Round type', 'Grade', 'Game date', 'Venue', 'Subvenue',
               'Time', 'Pool', 'Team 1 ', 'Team 1 score', 'Team 2', 'Team 2 score',
               'Duty team', 'Label']
    df = pd.DataFrame(all_rows, columns=columns)
    df["IS_BYE"] = df["Team 2"] == "BYE"
    df.sort_values(by=["Round", "IS_BYE", "Grade"], inplace=True)
    df.drop(columns="IS_BYE", inplace=True)
    df.to_csv(output_path, index=False)
    print(f"Rev format schedule exported to {output_path}")


# ============== Slot Analysis ==============

class SlotAnalyzer:
    """Analyze timeslot usage and find available slots."""
    
    def __init__(self, draw: DrawStorage, data: Dict):
        self.draw = draw
        self.data = data
        self.timeslots = data.get('timeslots', [])
    
    def get_all_possible_slots(self) -> List[Dict]:
        """Get all possible timeslots from the configuration."""
        slots = []
        for t in self.timeslots:
            slots.append({
                'week': t.week,
                'day': t.day,
                'day_slot': t.day_slot,
                'time': t.time,
                'date': t.date,
                'round_no': t.round_no,
                'field_name': t.field.name,
                'field_location': t.field.location
            })
        return slots
    
    def get_used_slots(self, week: Optional[int] = None) -> List[Dict]:
        """Get slots currently used by scheduled games."""
        used = []
        for game in self.draw.games:
            if week is not None and game.week != week:
                continue
            used.append({
                'week': game.week,
                'day': game.day,
                'day_slot': game.day_slot,
                'time': game.time,
                'date': game.date,
                'field_name': game.field_name,
                'field_location': game.field_location,
                'game_id': game.game_id,
                'teams': f"{game.team1} vs {game.team2}",
                'grade': game.grade
            })
        return used
    
    def get_unused_slots(self, week: Optional[int] = None) -> List[Dict]:
        """
        Get unused timeslots (available for scheduling).
        
        Returns slots that exist in timeslots config but have no game.
        """
        all_slots = self.get_all_possible_slots()
        if week is not None:
            all_slots = [s for s in all_slots if s['week'] == week]
        
        # Build set of used slot keys
        used_keys = set()
        for game in self.draw.games:
            if week is not None and game.week != week:
                continue
            key = (game.week, game.day_slot, game.field_name)
            used_keys.add(key)
        
        # Filter to unused
        unused = []
        for slot in all_slots:
            key = (slot['week'], slot['day_slot'], slot['field_name'])
            if key not in used_keys:
                unused.append(slot)
        
        return unused
    
    def slot_usage_summary(self) -> pd.DataFrame:
        """
        Generate summary of slot usage by week and field.
        
        Returns DataFrame with utilization percentages.
        """
        weeks = sorted(set(t.week for t in self.timeslots))
        fields = sorted(set(t.field.name for t in self.timeslots))
        
        data_rows = []
        for week in weeks:
            row = {'Week': week}
            week_total_possible = 0
            week_used = 0
            
            for field in fields:
                possible = len([t for t in self.timeslots 
                               if t.week == week and t.field.name == field])
                used = len([g for g in self.draw.games 
                           if g.week == week and g.field_name == field])
                week_total_possible += possible
                week_used += used
                
                if possible > 0:
                    row[field] = f"{used}/{possible}"
                else:
                    row[field] = "-"
            
            if week_total_possible > 0:
                row['Utilization'] = f"{week_used/week_total_possible*100:.1f}%"
            else:
                row['Utilization'] = "-"
            data_rows.append(row)
        
        return pd.DataFrame(data_rows)
    
    def print_unused_slots(self, week: Optional[int] = None) -> None:
        """Print unused slots to console."""
        unused = self.get_unused_slots(week)
        
        if not unused:
            print(f"No unused slots found" + (f" in week {week}" if week else ""))
            return
        
        print(f"\n{'='*60}")
        print(f"UNUSED SLOTS" + (f" - Week {week}" if week else " - All Weeks"))
        print(f"{'='*60}")
        
        # Group by week
        by_week = defaultdict(list)
        for slot in unused:
            by_week[slot['week']].append(slot)
        
        for wk in sorted(by_week.keys()):
            print(f"\nWeek {wk}:")
            for slot in sorted(by_week[wk], key=lambda x: (x['day_slot'], x['field_name'])):
                print(f"  Slot {slot['day_slot']}: {slot['field_name']} @ {slot['field_location']} "
                      f"({slot['day']} {slot['time']})")
        
        print(f"\nTotal unused: {len(unused)} slots")


def get_slot_analyzer(draw: DrawStorage, data: Dict) -> SlotAnalyzer:
    """Create a slot analyzer for a draw."""
    return SlotAnalyzer(draw, data)
