# tests/test_export_schedule_xlsx_notes.py
"""
No-mock tests for the ``weekend_notes`` column in
``DrawStorage.export_schedule_xlsx`` (spec-028 Unit B).

Three Given/When/Then scenarios with hand-computed oracles verified against
the real export row layout.  No mocks or patches — DrawStorage is built
synthetically and the written xlsx is re-opened with openpyxl.

Observed row layout (verified by probe):
  - Row 1  : week title  (col 1 = "Week N")
  - Row 2  : column headers  (col 14 = "Notes" when weekend_notes supplied)
  - Row 3  : field sub-header  ("EF - Newcastle International Hockey Centre")
  - Rows 4+ : game data rows  (first_game_row = 4 for a single-field week)

Column index map:
  A=1..K=11  game data
  L=12, M=13  blank spacers (never written)
  N=14  Notes column  (notes_col = NUM_COLS + 3 = 14)
"""

import os
import sys
import tempfile

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from analytics.storage import DrawStorage, StoredGame


# ---------------------------------------------------------------------------
# Shared helpers  (mirror test_weekend_notes.py pattern)
# ---------------------------------------------------------------------------

NIHC = "Newcastle International Hockey Centre"


def _make_game(
    game_id: str,
    date: str,
    week: int,
    time: str = "10:00",
    day_slot: int = 1,
    field_name: str = "EF",
    team1: str = "Norths PHL",
    team2: str = "Wests PHL",
    grade: str = "PHL",
    field_location: str = NIHC,
) -> StoredGame:
    """Return a minimal StoredGame."""
    return StoredGame(
        game_id=game_id,
        team1=team1,
        team2=team2,
        grade=grade,
        week=week,
        round_no=week,
        date=date,
        day="Sunday",
        time=time,
        day_slot=day_slot,
        field_name=field_name,
        field_location=field_location,
    )


def _make_draw(*games: StoredGame) -> DrawStorage:
    """Return a DrawStorage containing the supplied games."""
    return DrawStorage(
        description="synthetic test draw",
        num_weeks=len({g.week for g in games}),
        num_games=len(games),
        games=list(games),
    )


def _export(draw: DrawStorage, tmp_path, suffix: str = "", **kwargs) -> openpyxl.Workbook:
    """Export draw to a temp xlsx and return the opened workbook."""
    path = str(tmp_path / f"schedule{suffix}.xlsx")
    draw.export_schedule_xlsx(path, **kwargs)
    return openpyxl.load_workbook(path), path


# ---------------------------------------------------------------------------
# Scenario 1 — 3 games in week 1, 3 stacked notes, exact cell positions
# ---------------------------------------------------------------------------

class TestScenario1_ThreeGamesThreeNotes:
    """
    Given a synthetic DrawStorage with 2 weeks:
      - Week 1: 3 games on one field (EF, NIHC) at times 10:00, 11:30, 13:00
      - Week 2: 1 game
    and weekend_notes={1: ["Field: X", "Request: Y", "Preferred: Z"]},
    when export_schedule_xlsx is called,
    then re-opening the xlsx with openpyxl yields:

    Observed row layout (verified by probe):
      row1  = week 1 title  ("Week 1")
      row2  = column headers  (N2 == "Notes")
      row3  = field sub-header  ("EF - Newcastle International Hockey Centre")
      row4  = game 1 (time 10:00)  → first_game_row=4  → N4 == "Field: X"
      row5  = game 2 (time 11:30)                        N5 == "Request: Y"
      row6  = game 3 (time 13:00)                        N6 == "Preferred: Z"
      row7+ = blank / week 2 start

    L and M are None throughout.  DoD #4.
    """

    @pytest.fixture
    def draw(self):
        return _make_draw(
            _make_game("G00001", "2026-03-22", 1, time="10:00", day_slot=1),
            _make_game("G00002", "2026-03-22", 1, time="11:30", day_slot=2,
                       team1="Maitland PHL", team2="Tigers PHL"),
            _make_game("G00003", "2026-03-22", 1, time="13:00", day_slot=3,
                       team1="Souths PHL", team2="Gosford PHL"),
            _make_game("G00004", "2026-03-29", 2, time="10:00", day_slot=1),
        )

    @pytest.fixture
    def wb(self, draw, tmp_path):
        wb, _ = _export(
            draw,
            tmp_path,
            weekend_notes={1: ["Field: X", "Request: Y", "Preferred: Z"]},
        )
        yield wb
        wb.close()

    def test_notes_header_in_col_N_row2(self, wb):
        # Oracle: header row is row 2 (row 1 = week title)
        ws = wb.active
        assert ws.cell(row=2, column=14).value == "Notes"

    def test_first_note_at_first_game_row(self, wb):
        # Oracle: first_game_row = 4 (title=1, header=2, field-sub=3, game1=4)
        ws = wb.active
        assert ws.cell(row=4, column=14).value == "Field: X"

    def test_second_note_one_row_below(self, wb):
        ws = wb.active
        assert ws.cell(row=5, column=14).value == "Request: Y"

    def test_third_note_two_rows_below(self, wb):
        ws = wb.active
        assert ws.cell(row=6, column=14).value == "Preferred: Z"

    def test_col_L_is_none_at_game_rows(self, wb):
        ws = wb.active
        # L = column 12 — must be None at game rows (4–6)
        for r in [4, 5, 6]:
            assert ws.cell(row=r, column=12).value is None, (
                f"Expected L (col 12) row {r} to be None"
            )

    def test_col_M_is_none_at_game_rows(self, wb):
        ws = wb.active
        # M = column 13 — must be None at game rows (4–6)
        for r in [4, 5, 6]:
            assert ws.cell(row=r, column=13).value is None, (
                f"Expected M (col 13) row {r} to be None"
            )

    def test_week1_title_row(self, wb):
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Week 1"

    def test_field_subheader_row(self, wb):
        ws = wb.active
        assert ws.cell(row=3, column=1).value == f"EF - {NIHC}"

    def test_game_data_in_cols_A_to_K(self, wb):
        ws = wb.active
        # First game row (row 4) must have game data in cols A-K
        assert ws.cell(row=4, column=1).value == 1   # week
        assert ws.cell(row=4, column=7).value == "Norths PHL"   # team1
        assert ws.cell(row=4, column=8).value == "Wests PHL"    # team2

    def test_week2_also_has_notes_header(self, wb):
        # Week 2 starts at row 9 (rows 7-8 are blank/separator after week 1).
        # Oracle: after week 1's 3 game rows (rows 4-6), blank at 7, blank at 8
        # (no byes — all teams played in week 1 from this draw), week 2 title=9.
        # Actually: blank row (row+=1 before byes) + no byes + gap row (row+=1)
        # -> row7 blank, row8 blank -> week2 title = row9.
        ws = wb.active
        assert ws.cell(row=9, column=1).value == "Week 2"
        assert ws.cell(row=10, column=14).value == "Notes"


# ---------------------------------------------------------------------------
# Scenario 2 — weekend_notes=None is byte-for-cell-identical in A-K
# ---------------------------------------------------------------------------

class TestScenario2_NoneNotes_AdditiveOnly:
    """
    Given the same draw as Scenario 1, when exported with weekend_notes=None
    (or omitted), then:
      - Every cell in columns L (12), M (13), N (14) is None for all rows.
      - Columns A–K match a baseline export that also omits weekend_notes
        (i.e. the feature is strictly additive and opt-in).

    Proves DoD #5.
    """

    @pytest.fixture
    def draw(self):
        return _make_draw(
            _make_game("G00001", "2026-03-22", 1, time="10:00", day_slot=1),
            _make_game("G00002", "2026-03-22", 1, time="11:30", day_slot=2,
                       team1="Maitland PHL", team2="Tigers PHL"),
            _make_game("G00003", "2026-03-22", 1, time="13:00", day_slot=3,
                       team1="Souths PHL", team2="Gosford PHL"),
            _make_game("G00004", "2026-03-29", 2, time="10:00", day_slot=1),
        )

    def test_all_L_M_N_none_when_no_notes(self, draw, tmp_path):
        # When: export with weekend_notes=None (default)
        wb, _ = _export(draw, tmp_path, suffix="_nonotes", weekend_notes=None)
        ws = wb.active
        max_row = ws.max_row
        non_none = []
        for r in range(1, max_row + 1):
            for col in [12, 13, 14]:  # L, M, N
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    non_none.append(f"row{r} col{col}: {v!r}")
        wb.close()
        assert non_none == [], (
            f"Expected no values in L/M/N when weekend_notes=None. "
            f"Found: {non_none}"
        )

    def test_A_to_K_match_baseline(self, draw, tmp_path):
        # Export twice: once with notes, once without.
        # A-K must be identical regardless of weekend_notes value.
        wb_with, _ = _export(
            draw, tmp_path, suffix="_with",
            weekend_notes={1: ["Field: X", "Request: Y", "Preferred: Z"]},
        )
        wb_without, _ = _export(
            draw, tmp_path, suffix="_without",
            weekend_notes=None,
        )
        ws_with = wb_with.active
        ws_without = wb_without.active

        max_row = max(ws_with.max_row, ws_without.max_row)
        diffs = []
        for r in range(1, max_row + 1):
            for col in range(1, 12):  # A-K = cols 1-11
                v_with = ws_with.cell(row=r, column=col).value
                v_without = ws_without.cell(row=r, column=col).value
                if v_with != v_without:
                    diffs.append(
                        f"row{r} col{col}: with_notes={v_with!r} "
                        f"without_notes={v_without!r}"
                    )
        wb_with.close()
        wb_without.close()

        assert diffs == [], (
            f"Columns A-K differ between notes and no-notes exports. "
            f"First differences: {diffs[:5]}"
        )

    def test_no_notes_header_when_none(self, draw, tmp_path):
        wb, _ = _export(draw, tmp_path, suffix="_hdr", weekend_notes=None)
        ws = wb.active
        # Column N header cell (row 2) must be None when weekend_notes is None
        assert ws.cell(row=2, column=14).value is None
        wb.close()


# ---------------------------------------------------------------------------
# Scenario 3 — more notes than game rows; overflow stacks downward into byes area
# ---------------------------------------------------------------------------

class TestScenario3_MoreNotesThanGameRows:
    """
    Given a week with 1 game and 4 notes, when exported, all 4 notes appear
    in consecutive N cells from first_game_row (=4) through first_game_row+3 (=7).
    Bye-row column A still holds bye text (N may hold overflow note text).

    Observed row layout (verified by probe):
      row1  = week 1 title
      row2  = column headers (N2 = "Notes")
      row3  = field sub-header
      row4  = 1 game row  (first_game_row=4)  → N4 = "Note A"
      row5  = (empty game area)               → N5 = "Note B"
      row6  = BYES label  (col1="BYES")       → N6 = "Note C"
      row7  = bye team    (col1="Maitland PHL")→ N7 = "Note D"
      row8  = bye team    (col1="Tigers PHL")  → N8 = None

    This confirms that column-N notes never overwrite A-K game data or
    column-A bye content — they occupy only column N.  DoD #4 / spec §note-placement.
    """

    @pytest.fixture
    def draw(self):
        # 4 teams in the draw (Norths, Wests, Maitland, Tigers).
        # Week 1 has only 1 game (Norths vs Wests) → Maitland and Tigers are byes.
        # Week 2 has a game for all 4 so they appear in the full team set.
        return _make_draw(
            _make_game("G00001", "2026-03-22", 1, time="10:00", day_slot=1),
            _make_game("G00002", "2026-03-29", 2, time="10:00", day_slot=1,
                       team1="Maitland PHL", team2="Tigers PHL"),
        )

    @pytest.fixture
    def wb(self, draw, tmp_path):
        wb, _ = _export(
            draw,
            tmp_path,
            weekend_notes={1: ["Note A", "Note B", "Note C", "Note D"]},
        )
        yield wb
        wb.close()

    def test_all_four_notes_present(self, wb):
        """All 4 notes appear in N4, N5, N6, N7 (first_game_row=4)."""
        ws = wb.active
        # Oracle: first_game_row = 4 (title=1, header=2, field-sub=3, game=4)
        assert ws.cell(row=4, column=14).value == "Note A"
        assert ws.cell(row=5, column=14).value == "Note B"
        assert ws.cell(row=6, column=14).value == "Note C"
        assert ws.cell(row=7, column=14).value == "Note D"

    def test_bye_row_col_A_unchanged(self, wb):
        """Column A of the BYES row still holds 'BYES' (N has a note, but A is untouched)."""
        ws = wb.active
        # Oracle: row6 col1 = "BYES" (verified by probe)
        assert ws.cell(row=6, column=1).value == "BYES"

    def test_bye_team_col_A_unchanged(self, wb):
        """Column A of a bye-team row still holds the team name."""
        ws = wb.active
        # Oracle: row7 col1 = "Maitland PHL" (first bye team, alphabetical)
        assert ws.cell(row=7, column=1).value == "Maitland PHL"

    def test_game_row_A_to_K_unaffected(self, wb):
        """The single game row (row 4) has correct A-K data AND N4 note."""
        ws = wb.active
        # col1=week(1), col7=team1, col8=team2
        assert ws.cell(row=4, column=1).value == 1
        assert ws.cell(row=4, column=7).value == "Norths PHL"
        assert ws.cell(row=4, column=8).value == "Wests PHL"
        # Note in N is also present
        assert ws.cell(row=4, column=14).value == "Note A"

    def test_no_note_in_col_L_or_M(self, wb):
        """Columns L (12) and M (13) are never written, even in overflow rows."""
        ws = wb.active
        for r in range(4, 9):
            assert ws.cell(row=r, column=12).value is None, f"L not None at row {r}"
            assert ws.cell(row=r, column=13).value is None, f"M not None at row {r}"
